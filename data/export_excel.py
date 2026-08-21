# -*- coding: utf-8 -*-
"""Export dataset Excel (7 event) → workbook multi-sheet."""
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "War_Tiket_Konser_Korea_Data.xlsx"
OUT_DL = Path(r"C:\Users\User\Downloads\War_Tiket_Konser_Korea_Data.xlsx")

header_fill = PatternFill("solid", fgColor="0F172A")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
title_font = Font(bold=True, name="Calibri", size=14, color="0F172A")
section_font = Font(bold=True, name="Calibri", size=12)
thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
fills = {
    1: PatternFill("solid", fgColor="EDE9FE"),  # TREASURE
    2: PatternFill("solid", fgColor="DBEAFE"),  # LYKN
    3: PatternFill("solid", fgColor="FCE7F3"),  # BLACKPINK
    4: PatternFill("solid", fgColor="FEF3C7"),  # NCT DREAM
    5: PatternFill("solid", fgColor="DCFCE7"),  # EXO
    6: PatternFill("solid", fgColor="FFEDD5"),  # ATEEZ
    7: PatternFill("solid", fgColor="E0E7FF"),  # BUS
}
wrap = Alignment(wrap_text=True, vertical="center")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_rows(ws, start, headers, rows, row_fill_fn=None):
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header(ws, start, len(headers))
    for ri, row in enumerate(rows):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=start + 1 + ri, column=ci, value=v)
            cell.border = thin
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=10)
            if row_fill_fn:
                cell.fill = row_fill_fn(row, ri)
    return start + 1 + len(rows)


def main():
    events = json.loads((ROOT / "events.manual.json").read_text(encoding="utf-8"))
    seats_path = ROOT / "seats.manual.csv"
    if not seats_path.exists():
        raise SystemExit("seats.manual.csv missing — run: npm run data:generate")

    # parse CSV with quotes
    seat_rows = []
    with seats_path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            seat_rows.append(r)

    wb = Workbook()

    # ---- 00 Panduan ----
    ws = wb.active
    ws.title = "00_Panduan"
    ws["A1"] = "DATA REAL — War Tiket Konser (Excel dataset)"
    ws["A1"].font = title_font
    ws["A2"] = "Squad Klpk1 · Sumber: DATA_WAR_TIKET_KONSER.xlsx · Total 2.480 kursi · 7 event"
    ws["A2"].font = Font(italic=True, color="475569")
    lines = [
        "",
        "Isi workbook:",
        "01_Events — 7 konser (TREASURE, LYKN, BLACKPINK, NCT DREAM, EXO, ATEEZ, BUS)",
        "02_Categories — zona/harga/kuota per event",
        "03_Seats — denah lengkap semua seat_code",
        "04_Summary — cek seats = quota_total",
        "05_Terms — syarat tiket per event",
        "",
        "Cara load ke sistem:",
        "  npm run data:excel",
        "  npm start  → http://localhost:3000/?event=1",
        "",
        "Catatan harga: IDR untuk demo praktikum (asli biasanya KRW).",
        "Anti-oversell: kuota dipotong atomik di Redis saat POST /orders — bukan di Excel.",
    ]
    for i, t in enumerate(lines):
        ws.cell(row=3 + i, column=1, value=t)
    autosize(ws, [90])

    # ---- 01 Events ----
    ws = wb.create_sheet("01_Events")
    ws["A1"] = "01 — Events (katalog konser Excel)"
    ws["A1"].font = title_font
    eheaders = [
        "event_id",
        "title",
        "artist",
        "venue",
        "city",
        "country",
        "starts_at",
        "ends_at",
        "sales_opens_at",
        "sales_closes_at",
        "quota_total",
        "price_idr_default",
        "status",
        "gate_open",
        "age_rating",
        "timezone",
        "description",
    ]
    erows = []
    for e in events:
        erows.append(
            [
                e["event_id"],
                e["title"],
                e["artist"],
                e["venue"],
                e.get("city"),
                e.get("country"),
                e.get("starts_at"),
                e.get("ends_at"),
                e.get("sales_opens_at"),
                e.get("sales_closes_at"),
                e.get("quota_total"),
                e.get("price_idr"),
                e.get("status"),
                e.get("gate_open"),
                e.get("age_rating"),
                e.get("timezone"),
                e.get("description"),
            ]
        )

    def efill(row, _ri):
        return fills.get(row[0], PatternFill())

    write_rows(ws, 3, eheaders, erows, efill)
    autosize(ws, [10, 45, 22, 50, 12, 14, 22, 22, 22, 22, 12, 14, 12, 12, 28, 14, 55])
    ws.row_dimensions[3].height = 22
    for r in range(4, 4 + len(erows)):
        ws.row_dimensions[r].height = 48

    # ---- 02 Categories ----
    ws = wb.create_sheet("02_Categories")
    ws["A1"] = "02 — Seat categories / zona harga"
    ws["A1"].font = title_font
    cheaders = [
        "event_id",
        "artist",
        "code",
        "name",
        "price_idr",
        "quota",
        "color_hex",
        "pct_of_event",
    ]
    crows = []
    for e in events:
        for c in e.get("categories", []):
            q = int(c["quota"])
            tot = int(e["quota_total"])
            crows.append(
                [
                    e["event_id"],
                    e["artist"],
                    c["code"],
                    c["name"],
                    c["price_idr"],
                    q,
                    c.get("color", ""),
                    round(100 * q / tot, 1) if tot else 0,
                ]
            )

    def cfill(row, _ri):
        return fills.get(row[0], PatternFill())

    write_rows(ws, 3, cheaders, crows, cfill)
    autosize(ws, [10, 22, 10, 40, 12, 10, 12, 12])

    # ---- 03 Seats (full) ----
    ws = wb.create_sheet("03_Seats")
    ws["A1"] = f"03 — Denah kursi lengkap ({len(seat_rows)} baris)"
    ws["A1"].font = title_font
    sheaders = [
        "event_id",
        "seat_code",
        "category",
        "category_name",
        "row_label",
        "seat_number",
        "section",
        "price_idr",
        "color_hex",
        "pos_x",
        "pos_y",
    ]
    srows = []
    for r in seat_rows:
        srows.append(
            [
                int(r["event_id"]) if r.get("event_id") else "",
                r.get("seat_code", ""),
                r.get("category", ""),
                r.get("category_name", ""),
                r.get("row_label", ""),
                int(r["seat_number"]) if str(r.get("seat_number", "")).isdigit() else r.get("seat_number"),
                r.get("section", ""),
                int(r["price_idr"]) if str(r.get("price_idr", "")).isdigit() else r.get("price_idr"),
                r.get("color_hex", ""),
                r.get("pos_x", ""),
                r.get("pos_y", ""),
            ]
        )

    def sfill(row, _ri):
        return fills.get(row[0], PatternFill())

    write_rows(ws, 3, sheaders, srows, sfill)
    autosize(ws, [10, 12, 10, 36, 10, 12, 32, 12, 12, 8, 8])
    ws.auto_filter.ref = f"A3:K{3 + len(srows)}"
    ws.freeze_panes = "A4"

    # ---- 04 Summary ----
    ws = wb.create_sheet("04_Summary")
    ws["A1"] = "04 — Ringkasan & validasi quota"
    ws["A1"].font = title_font
    # count seats from CSV
    from collections import Counter, defaultdict

    by_ev = Counter()
    by_ev_cat = defaultdict(Counter)
    for r in seat_rows:
        eid = int(r["event_id"])
        by_ev[eid] += 1
        by_ev_cat[eid][r["category"]] += 1

    sum_headers = [
        "event_id",
        "artist",
        "city",
        "quota_total_json",
        "seats_in_csv",
        "match",
        "breakdown_kategori",
    ]
    sum_rows = []
    for e in events:
        eid = e["event_id"]
        csv_n = by_ev[eid]
        match = "YES" if csv_n == int(e["quota_total"]) else "NO"
        br = ", ".join(f"{k}={v}" for k, v in sorted(by_ev_cat[eid].items()))
        sum_rows.append(
            [
                eid,
                e["artist"],
                e.get("city"),
                e["quota_total"],
                csv_n,
                match,
                br,
            ]
        )
    write_rows(ws, 3, sum_headers, sum_rows, lambda row, _ri: fills.get(row[0], PatternFill()))
    total_seats = sum(by_ev.values())
    ws.cell(row=4 + len(sum_rows) + 1, column=1, value=f"TOTAL SEATS CSV = {total_seats}").font = section_font
    autosize(ws, [10, 24, 12, 16, 14, 10, 50])

    # ---- 05 Terms ----
    ws = wb.create_sheet("05_Terms")
    ws["A1"] = "05 — Syarat & ketentuan per event"
    ws["A1"].font = title_font
    theaders = ["event_id", "artist", "no", "term"]
    trows = []
    for e in events:
        for i, t in enumerate(e.get("terms", []), 1):
            trows.append([e["event_id"], e["artist"], i, t])
    write_rows(ws, 3, theaders, trows, lambda row, _ri: fills.get(row[0], PatternFill()))
    autosize(ws, [10, 24, 6, 70])

    # ---- 06 How to use ----
    ws = wb.create_sheet("06_Cara_Pakai")
    ws["A1"] = "06 — Cara pakai Excel ini di proyek"
    ws["A1"].font = title_font
    cara = [
        "1. Sheet 01–03 = sumber data master (boleh diedit di Excel).",
        "2. Setelah edit Events/Categories: sesuaikan data/events.manual.json (atau minta Data Engineer sync).",
        "3. Setelah edit denah besar: lebih aman edit generator data/generate-real-seats.js lalu npm run data:generate.",
        "4. Load ke DB: npm run data:excel  (atau npm run data:reload).",
        "5. Web: http://localhost:3000/ — event 1–7 (TREASURE … BUS).",
        "6. Filter sheet 03_Seats pakai AutoFilter (baris header) → filter event_id atau category.",
        "7. Jangan mengandalkan Excel untuk cegah oversell — itu tugas Redis DECR di backend.",
    ]
    for i, t in enumerate(cara):
        ws.cell(row=3 + i, column=1, value=t)
    autosize(ws, [100])

    wb.save(OUT)
    wb.save(OUT_DL)
    print("SAVED", OUT)
    print("SAVED", OUT_DL)
    print("events", len(events), "seats", len(seat_rows), "categories", len(crows))


if __name__ == "__main__":
    main()
